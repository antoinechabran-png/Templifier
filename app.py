import streamlit as st
import pandas as pd
import io
import re
import openpyxl
from openpyxl.utils import get_column_letter

# Set page config
st.set_page_config(page_title="Market Research Templifier", layout="wide")

# Constants
GENERAL_METRICS = ["Mean", "Top Box", "Top 2 Boxes", "Top 3 Boxes", "Bottom Box", "Bottom 2 Boxes", "Bottom 3 Boxes"]
PASTELS = ['#E3F2FD', '#F3E5F5', '#E8F5E9', '#FFF3E0', '#FCE4EC'] # Soft Blue, Purple, Green, Orange, Pink
SOFT_BORDER = '#B0BEC5' # Muted blue-grey

def get_question_root(q_text):
    match = re.match(r"^(Q-\d+|S-\d+)", str(q_text))
    return match.group(1) if match else str(q_text)

st.title("📊 Market Research Templifier")

uploaded_file = st.file_uploader("Upload Raw Results (Excel)", type=["xlsx"])

if uploaded_file is not None:
    # --- 1. SETTINGS & LOADING ---
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheet_names = wb.sheetnames
    selected_sheet = st.selectbox("Select the main sheet to process", sheet_names)
    ws = wb[selected_sheet]
    
    # Delta Integration Settings
    st.sidebar.header("Integration Settings")
    use_deltas = st.sidebar.toggle("Enable Delta Integration", value=False)
    df_delta = None
    delta_metadata = {}
    if use_deltas:
        delta_sheet_name = st.sidebar.selectbox("Select the Delta sheet", sheet_names)
        ws_delta = wb[delta_sheet_name]
        df_delta = pd.DataFrame(ws_delta.values)
        # Capture metadata for the Delta sheet to identify percentage cells
        for r in range(1, ws_delta.max_row + 1):
            for c in range(1, ws_delta.max_column + 1):
                cell = ws_delta.cell(row=r, column=c)
                if cell.number_format and '%' in cell.number_format:
                    delta_metadata[(r-1, c-1)] = True

    # Sidebar Data Structure Settings
    num_benchmarks = st.sidebar.slider("Number of Benchmarks in Raw File", 1, 5, 2)
    bench_start_col = 2
    product_start_col = bench_start_col + (num_benchmarks * 2)
    cols_per_product = 2 + num_benchmarks

    # Load main data into DataFrame
    df = pd.DataFrame(ws.values)
    # Cleanup Question and Metric columns
    df[0] = df[0].astype(str).str.strip()
    df[1] = df[1].astype(str).str.strip()

    # Capture metadata (Color and Percentages) from the main sheet
    cell_metadata = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            meta = {"color": None, "is_percent": False}
            if cell.fill and cell.fill.start_color.index != "00000000" and cell.fill.start_color.rgb != "FFFFFFFF":
                color_hex = cell.fill.start_color.rgb
                if isinstance(color_hex, str) and len(color_hex) == 8:
                    meta["color"] = f"#{color_hex[2:]}"
            if cell.number_format and '%' in cell.number_format:
                meta["is_percent"] = True
            if meta["color"] or meta["is_percent"]:
                cell_metadata[(r-1, c-1)] = meta

    # --- 2. PARSING PRODUCTS ---
    product_triplets = {} 
    for col_idx in range(product_start_col, len(df.columns) - (cols_per_product - 1), cols_per_product):
        p_name = df.iloc[2, col_idx]
        if pd.isna(p_name) or str(p_name).strip().lower() in ["none", "nan", ""]: 
            p_name = f"Product at Col {get_column_letter(col_idx+1)}"
        product_triplets[str(p_name).strip()] = list(range(col_idx, col_idx + cols_per_product))

    # --- SIDEBAR FILTERS ---
    regroup_mode = st.sidebar.toggle("Enable Question Regrouping", value=True)
    all_p_names = list(product_triplets.keys())
    selected_products = st.sidebar.multiselect("Select Products to Include", all_p_names, default=all_p_names)
    show_sig = st.sidebar.checkbox("Show Significance/Delta Columns", value=True)

    # --- UI FILTERING (METRIC SELECTION) ---
    st.header("🎯 Question & Metric Selection")
    # Identify unique questions
    raw_data_area = df.iloc[5:, [0, 1]].dropna(subset=[0])
    ui_q_map = {}
    for q_full in raw_data_area[0].unique():
        if q_full in ["nan", "None", ""]: continue
        display_name = get_question_root(q_full) if regroup_mode else q_full
        metrics = df[df[0] == q_full][1].unique().tolist()
        if display_name not in ui_q_map:
            ui_q_map[display_name] = {"originals": [q_full], "metrics": set(metrics)}
        else:
            ui_q_map[display_name]["originals"].append(q_full)
            ui_q_map[display_name]["metrics"].update(metrics)

    selected_q_metrics = {}
    for display_q, data in ui_q_map.items():
        with st.expander(f"❓ {display_q}"):
            metrics_list = sorted([str(m) for m in data["metrics"] if pd.notna(m)])
            gen_group = [m for m in metrics_list if m in GENERAL_METRICS]
            mod_group = [m for m in metrics_list if m not in GENERAL_METRICS]
            
            c1, c2, c3 = st.columns([1, 2, 2])
            is_active = c1.checkbox("Keep Question", value=True, key=f"act_{display_q}")
            
            if is_active:
                m_gen_key, m_mod_key = f"mg_{display_q}", f"mm_{display_q}"
                
                # Default selection (Select All) logic
                if m_gen_key not in st.session_state: st.session_state[m_gen_key] = gen_group
                if m_mod_key not in st.session_state: st.session_state[m_mod_key] = mod_group
                
                with c2:
                    st.write("**General Metrics**")
                    col_a, col_b = st.columns(2)
                    if col_a.button("All", key=f"allg_{display_q}"): 
                        st.session_state[m_gen_key] = gen_group
                        st.rerun()
                    if col_b.button("None", key=f"clrg_{display_q}"): 
                        st.session_state[m_gen_key] = []
                        st.rerun()
                    sel_gen = st.multiselect("Select", gen_group, key=m_gen_key)
                    
                with c3:
                    st.write("**Modalities**")
                    col_a, col_b = st.columns(2)
                    if col_a.button("All", key=f"allm_{display_q}"): 
                        st.session_state[m_mod_key] = mod_group
                        st.rerun()
                    if col_b.button("None", key=f"clrm_{display_q}"): 
                        st.session_state[m_mod_key] = []
                        st.rerun()
                    sel_mod = st.multiselect("Select", mod_group, key=m_mod_key)
                
                # Apply selection to all original sub-questions
                for orig in data["originals"]:
                    selected_q_metrics[orig] = set(sel_gen + sel_mod)

    # --- 3. EXPORT WITH DELTA INTEGRATION ---
    if st.button("🚀 Generate Beautiful Excel"):
        data_rows = []
        for idx, row in df.iloc[5:].iterrows():
            q_name, m_name = str(row[0]).strip(), str(row[1]).strip()
            if q_name in selected_q_metrics and m_name in selected_q_metrics[q_name]:
                data_rows.append((idx, row))

        if not data_rows:
            st.error("No questions selected. Please select at least one question and metric.")
        else:
            cols_to_keep = [0, 1]
            benchmark_cols = list(range(bench_start_col, product_start_col))
            cols_to_keep.extend(benchmark_cols)
            
            final_product_cols = []
            for p in selected_products:
                indices = product_triplets[p]
                # If show_sig is true, keep Value + Sig + Deltas. Otherwise just Value.
                kept = indices if show_sig else indices[:1]
                final_product_cols.append({'name': p, 'indices': kept})
                cols_to_keep.extend(kept)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                # Prepare base rows (header rows 0-4 + filtered data rows)
                final_rows = [df.iloc[i].copy() for i in range(5)] + [r[1].copy() for r in data_rows]
                
                # Delta Integration Logic: Override values in delta columns from the delta sheet
                if use_deltas and df_delta is not None:
                    for row_idx, row in enumerate(final_rows):
                        # Map current row back to the original source index
                        source_row_idx = row_idx if row_idx < 5 else data_rows[row_idx - 5][0]
                        for p_info in final_product_cols:
                            # Delta columns are the 3rd column onwards in each product block (indices[2:])
                            delta_idxs = p_info['indices'][2:]
                            for col_idx in delta_idxs:
                                if col_idx < df_delta.shape[1]:
                                    raw_val = df_delta.iloc[source_row_idx, col_idx]
                                    # If it was a % in the delta sheet, convert to a 'whole' number for display
                                    if delta_metadata.get((source_row_idx, col_idx), False) and isinstance(raw_val, (int, float)):
                                        row.iloc[col_idx] = raw_val * 100
                                    else:
                                        row.iloc[col_idx] = raw_val

                final_df = pd.DataFrame(final_rows)[cols_to_keep].reset_index(drop=True)
                final_df.to_excel(writer, index=False, header=False, sheet_name='Report')
                
                workbook = writer.book
                worksheet = writer.sheets['Report']
                worksheet.freeze_panes(5, 2)

                # Formats
                header_fmt = workbook.add_format({'bold': True, 'bg_color': '#455A64', 'font_color': 'white', 'border': 1, 'border_color': SOFT_BORDER, 'align': 'center', 'valign': 'vcenter'})
                sub_header_fmt = workbook.add_format({'bold': True, 'bg_color': '#CFD8DC', 'border': 1, 'border_color': SOFT_BORDER, 'align': 'center'})
                pastel_fmts = [workbook.add_format({'bg_color': c, 'border': 1, 'border_color': SOFT_BORDER, 'valign': 'vcenter', 'text_wrap': True}) for c in PASTELS]

                # A. Apply Headers
                for b_idx in range(num_benchmarks):
                    b_col = 2 + (b_idx * 2)
                    worksheet.merge_range(2, b_col, 2, b_col + 1, str(df.iloc[2, b_col]), header_fmt)

                curr_c = product_start_col
                for p_info in final_product_cols:
                    n_cols = len(p_info['indices'])
                    p_head_fmt = workbook.add_format({'bold': True, 'bg_color': '#455A64', 'font_color': 'white', 'border': 1, 'border_color': SOFT_BORDER, 'left': 2, 'right': 2, 'align': 'center', 'valign': 'vcenter'})
                    if n_cols > 1:
                        worksheet.merge_range(2, curr_c, 2, curr_c + n_cols - 1, p_info['name'], p_head_fmt)
                    else:
                        worksheet.write(2, curr_c, p_info['name'], p_head_fmt)
                    
                    # Apply sub-header labels from original Row 5
                    for i, col_idx in enumerate(p_info['indices']):
                        val = final_df.iloc[4, curr_c + i]
                        worksheet.write(4, curr_c + i, val if pd.notna(val) else "", sub_header_fmt)
                    curr_c += n_cols

                # B. Data Body Styling & Deltas
                start_r = 5
                pastel_idx = -1
                last_root = None
                for r in range(5, len(final_df)):
                    q_val = str(final_df.iloc[r, 0])
                    root = get_question_root(q_val)
                    is_last_text = (r == len(final_df) - 1 or str(final_df.iloc[r+1, 0]) != q_val)
                    
                    # Grouping Color logic
                    if is_last_text:
                        if root != last_root:
                            pastel_idx += 1
                            last_root = root
                        fmt = pastel_fmts[pastel_idx % len(PASTELS)]
                        if r > start_r: worksheet.merge_range(start_r, 0, r, 0, q_val, fmt)
                        else: worksheet.write(start_r, 0, q_val, fmt)
                        start_r = r + 1

                    orig_row_idx = data_rows[r - 5][0]
                    
                    for target_c in range(1, len(final_df.columns)):
                        orig_col_idx = cols_to_keep[target_c]
                        val = final_df.iloc[r, target_c]
                        meta = cell_metadata.get((orig_row_idx, orig_col_idx), {"color": None, "is_percent": False})
                        
                        cell_style = {'border': 1, 'border_color': SOFT_BORDER, 'bottom': 1 if is_last_text else 4, 'align': 'center'}
                        
                        # Product block boundaries and delta identification
                        is_delta_col = False
                        is_color_col = False
                        if target_c >= product_start_col:
                            block_size = len(final_product_cols[0]['indices'])
                            pos_in_block = (target_c - product_start_col) % block_size
                            if pos_in_block == 0: cell_style['left'] = 2
                            if pos_in_block == block_size - 1: cell_style['right'] = 2
                            
                            if pos_in_block == 1: is_color_col = True # Sig Column
                            if pos_in_block >= 2: is_delta_col = True # Delta Columns

                        if is_color_col:
                            val = "" # Clear text for sig-color columns
                        
                        if meta["color"]: 
                            cell_style['bg_color'] = meta["color"]
                        
                        # Format numbers
                        if not is_delta_col:
                            # Standard percentages for Top Box/Mean
                            if meta["is_percent"] or (str(final_df.iloc[r, 1]) not in GENERAL_METRICS and isinstance(val, (int, float))):
                                cell_style['num_format'] = '0%'
                            elif isinstance(val, (float, int)) and abs(val) < 1 and val != 0:
                                cell_style['num_format'] = '0.0%'
                        else:
                            # Deltas: Display as plain numbers (no % sign)
                            cell_style['num_format'] = '0.00' if isinstance(val, float) else '0'

                        worksheet.write(r, target_c, val if pd.notna(val) else "", workbook.add_format(cell_style))

                worksheet.set_column(0, 0, 35)
                worksheet.set_column(1, 1, 20)

            st.success("✅ Report with integrated Deltas generated!")
            st.download_button("📥 Download Excel", output.getvalue(), f"Templified_Report.xlsx")
